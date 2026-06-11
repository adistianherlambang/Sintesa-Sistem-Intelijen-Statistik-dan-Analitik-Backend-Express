import openpyxl
import os

wb = openpyxl.load_workbook("hasilAnalisis.xlsx", data_only=True)
print("Sheet Names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ({sheet.max_row} rows x {sheet.max_column} cols) ---")
    for r in range(1, min(10, sheet.max_row + 1)):
        row_vals = [sheet.cell(r, c).value for c in range(1, min(8, sheet.max_column + 1))]
        print(f"Row {r}: {row_vals}")
